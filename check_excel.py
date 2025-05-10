from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

def inspect_excel():
    print("\n=== Excel File Inspector ===")
    wb = None
    
    try:
        # Load both formula and value versions
        wb_formulas = load_workbook('Copy of MASTER TEMPLATE.xlsx', data_only=False)
        wb_values = load_workbook('Copy of MASTER TEMPLATE.xlsx', data_only=True)
        ws_formulas = wb_formulas.active
        ws_values = wb_values.active
        
        # 1. Header Section Analysis
        print("\n=== Header Section ===")
        print("Circulator & Date (Row 2):")
        for col in ['A', 'B']:
            value = ws_values[f'{col}2'].value
            print(f"{col}2: {value}")
            
        print("\nSubheader Row (Row 3):")
        headers = ['Sub:', 'Price', '%Reg', 'Validity', 'Purge', 'Purge Rounded', 'Sum', 'Payable']
        for col, header in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], headers):
            value = ws_values[f'{col}3'].value
            print(f"{col}3 ({header}): {value}")
            
        print("\nValues Row (Row 4):")
        for col in range(1, 9):
            value = ws_values.cell(row=4, column=col).value
            print(f"{get_column_letter(col)}4: {value}")
            
        # 2. Coordinator Section
        print("\n=== Coordinator Section ===")
        print("Row 5:", ws_values['A5'].value)
        
        print("\nColumn Headers (Row 6):")
        headers = {
            'A': 'Petition Sheet',
            'J': 'Total Number of 1 & 0.1',
            'K': 'Dup [D]',
            'L': 'Purge [P]',
            'M': 'Total purge',
            'N': 'Count [C]',
            'O': 'Total Number of V',
            'P': 'Good [G]',
            'Q': 'Bad [B]'
        }
        for col, desc in headers.items():
            value = ws_values[f'{col}6'].value
            print(f"{col}6 ({desc}): {value}")
            
        # 3. Data Rows Analysis
        print("\n=== Data Rows Analysis ===")
        for row in range(7, 76):
            if any(ws_values.cell(row=row, column=col).value for col in range(1, 18)):
                print(f"\nRow {row}:")
                # Petition number
                print(f"Petition: {ws_values.cell(row=row, column=1).value}")
                
                # Signature cells (B-J)
                print("Signatures (B-J):", end=" ")
                for col in range(2, 11):
                    value = ws_values.cell(row=row, column=col).value
                    if value:
                        print(f"{get_column_letter(col)}{row}={value}", end=" ")
                
                # Calculations (K-Q)
                print("\nCalculations:")
                for col, desc in headers.items():
                    if col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
                        value = ws_values[f'{col}{row}'].value
                        if value:
                            print(f"{desc}: {value}")
                
        # Add Formula Analysis
        print("\n=== Formula Analysis ===")
        for row in range(7, 76):
            if any(ws_values.cell(row=row, column=col).value for col in range(1, 18)):
                print(f"\nRow {row} Analysis:")
                # Show both formulas and their results
                for col in range(10, 18):  # J through Q
                    formula = ws_formulas.cell(row=row, column=col).value
                    value = ws_values.cell(row=row, column=col).value
                    col_letter = get_column_letter(col)
                    print(f"{col_letter}{row}:")
                    print(f"  Formula: {formula}")
                    print(f"  Value: {value}")
                
        # 4. Footer Row Analysis
        print("\n=== Footer Row (76) ===")
        footer_row = 76
        print(f"A{footer_row}: {ws_values.cell(row=footer_row, column=1).value}")
        for col in range(10, 18):
            value = ws_values.cell(row=footer_row, column=col).value
            if value:
                print(f"{get_column_letter(col)}{footer_row}: {value}")
        
        # Add Footer Formula Analysis
        print("\n=== Footer Formula Analysis ===")
        row = 76
        for col in range(10, 18):
            formula = ws_formulas.cell(row=row, column=col).value
            value = ws_values.cell(row=row, column=col).value
            col_letter = get_column_letter(col)
            print(f"{col_letter}{row}:")
            print(f"  Formula: {formula}")
            print(f"  Value: {value}")
        
    except Exception as e:
        print(f"Error inspecting Excel: {e}")
    finally:
        if wb_formulas:
            wb_formulas.close()
        if wb_values:
            wb_values.close()

def fix_excel_template():
    wb = load_workbook('Copy of MASTER TEMPLATE.xlsx')
    ws = wb.active
    
    # Unmerge all cells first - fixed the iteration error
    merged_ranges = list(ws.merged_cells.ranges)  # Convert to list first
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    
    # Fix column headers (Row 6)
    headers = {
        'J6': 'Total Number of 1 & 0.1',
        'K6': 'Dup [D]',
        'L6': 'Purge [P]',
        'M6': 'Total purge = Dup + Purge',
        'N6': 'Count [C]',
        'O6': 'Checked [Ch]',
        'P6': 'Good [G]',
        'Q6': 'Bad [B]'
    }
    
    for cell, value in headers.items():
        ws[cell] = value
    
    # Fix formulas for all data rows
    for row in range(7, 76):
        formulas = {
            f'J{row}': f'=SUM(IF(B{row}:J{row}=1,1,IF(B{row}:J{row}=0.1,0.1,0)))',
            f'K{row}': f'=COUNTIF(B{row}:J{row},"x")',
            f'L{row}': f'=COUNTIF(B{row}:J{row},"v")',
            f'M{row}': f'=K{row}+L{row}',
            f'N{row}': f'=COUNTA(B{row}:J{row})',
            f'O{row}': f'=COUNTIF(B{row}:J{row},1)+COUNTIF(B{row}:J{row},0.1)',
            f'P{row}': f'=COUNTIF(B{row}:J{row},1)',
            f'Q{row}': f'=COUNTIF(B{row}:J{row},0.1)'
        }
        
        for col_letter, formula in formulas.items():
            try:
                cell = ws[col_letter]
                cell.value = formula
                cell.data_type = 'f'  # Set as formula
            except AttributeError:
                # If cell is merged, set the value to the top-left cell
                row_num = int(col_letter[1:])
                col = get_column_letter(ws[col_letter].column)
                ws[f'{col}{row_num}'] = formula
    
    # Fix header row values (Row 4)
    header_values = {
        'B4': 1.0,
        'C4': 0.1610305958,
        'D4': 61.53846154,
        'E4': "Total number of X= 2",
        'F4': 2,
        'G4': 18,
        'H4': 18.00
    }
    
    for cell, value in header_values.items():
        try:
            ws[cell] = value
        except AttributeError:
            pass  # Skip if cell is merged
    
    # Fix footer row (76)
    footer_formulas = {
        'J76': '="Total = "&SUM(J7:J75)',
        'K76': '="Total Dup = "&SUM(K7:K75)',
        'L76': '="Total Purge = "&SUM(L7:L75)',
        'M76': '="Total D+P = "&SUM(M7:M75)',
        'N76': '="Total Count = "&SUM(N7:N75)',
        'O76': '="Total Checked = "&SUM(O7:O75)',
        'P76': '="Total Good = "&SUM(P7:P75)',
        'Q76': '="Total Bad = "&SUM(Q7:Q75)'
    }
    
    for cell, formula in footer_formulas.items():
        try:
            ws[cell] = formula
            ws[cell].data_type = 'f'
        except AttributeError:
            pass  # Skip if cell is merged
    
    # Add data validation for signature cells (B-J)
    for row in range(7, 76):
        for col in range(2, 11):  # B through J
            try:
                cell = ws.cell(row=row, column=col)
                cell.number_format = '0.0'  # Format for numbers
            except AttributeError:
                pass  # Skip if cell is merged
    
    # Enable formula calculation
    wb.calculate_formulas = True
    
    wb.save('Copy of MASTER TEMPLATE.xlsx')
    print("Excel template has been fixed!")

if __name__ == "__main__":
    fix_excel_template()
    inspect_excel()  # Run inspection after fixing