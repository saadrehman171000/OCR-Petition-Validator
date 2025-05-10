from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

def inspect_petition(petition_number):
    print(f"\n=== Inspecting Petition {petition_number} ===")
    wb = None
    
    try:
        wb = load_workbook('Copy of MASTER TEMPLATE.xlsx', data_only=True)
        ws = wb.active
        
        # Find petition row
        row = None
        for r in range(7, 76):
            if ws.cell(row=r, column=1).value == petition_number:
                row = r
                break
        
        if row is None:
            print(f"Petition {petition_number} not found!")
            return
        
        # Show signature values
        print("\nSignature Values:")
        for col in range(2, 11):  # B through J
            value = ws.cell(row=row, column=col).value
            if value:
                print(f"{get_column_letter(col)}{row}: {value}")
        
        # Show calculations
        print("\nCalculations:")
        calcs = {
            'Total': ('J', 10),
            'Duplicates': ('K', 11),
            'Purges': ('L', 12),
            'Total Purge': ('M', 13),
            'Count': ('N', 14),
            'Checked': ('O', 15),
            'Good': ('P', 16),
            'Bad': ('Q', 17)
        }
        
        for desc, (col_letter, col_num) in calcs.items():
            value = ws.cell(row=row, column=col_num).value
            print(f"{desc} ({col_letter}{row}): {value}")
        
    except Exception as e:
        print(f"Error: {e}")
    finally:
        if wb:
            wb.close()

def list_all_petitions():
    print("\n=== All Petitions ===")
    wb = None
    
    try:
        wb = load_workbook('Copy of MASTER TEMPLATE.xlsx', data_only=True)
        ws = wb.active
        
        for row in range(7, 76):
            petition_num = ws.cell(row=row, column=1).value
            if petition_num:
                print(f"\nPetition {petition_num} (Row {row}):")
                # Show first few signatures
                print("Signatures:", end=" ")
                for col in range(2, 8):  # B through G
                    value = ws.cell(row=row, column=col).value
                    if value:
                        print(f"{get_column_letter(col)}{row}={value}", end=" ")
                print()
                
    except Exception as e:
        print(f"Error: {e}")
    finally:
        if wb:
            wb.close()

if __name__ == "__main__":
    # List all petitions
    list_all_petitions()
    
    # Ask which petition to inspect
    petition = input("\nEnter petition number to inspect (or press Enter to exit): ")
    if petition:
        inspect_petition(int(petition)) 