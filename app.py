# Standard library imports
import base64
import os
import io
from time import sleep
from datetime import datetime
import json
from bson.binary import Binary
from flask_cors import CORS
import pandas as pd
from datetime import datetime
from bson import ObjectId
from flask import Flask, request, jsonify, send_from_directory
from flask_socketio import SocketIO, emit

# Third-party imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bcrypt import hashpw, gensalt, checkpw
from flask_cors import CORS
from pymongo import MongoClient
from PIL import Image as PILImage
from google.cloud import vision
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import csv
import time

#
# from google.cloud.vision import types as vision_types

# Your custom imports
from automation import process_petition_text, authenticate
from automation_helper import send_details_by_xpath

# Local imports
from helpers import (
    google_cred_setup,
    rotate_image_based_on_metadata,
    concatenate_pil_images_vertically_with_marginssss,  # Corrected function name
    extract_info_updated,
    get_details_by_text,
    display_results_of_searching
)
from detection import ReceiptDetection
from fields import ReceiptFields
from constants import UPLOAD_FOLDER, PROCESSED_FOLDER

# Add at the top of app.py with other imports

from flask.json.provider import JSONProvider
from bson import ObjectId, binary

from config import JSON_DIR, CSV_DIR, EXCEL_DIR, ensure_data_directories

# Create Flask app first
app = Flask(__name__)

# Add CustomJSONProvider class back
class CustomJSONProvider(JSONProvider):
    def dumps(self, obj, **kwargs):
        def custom_default(obj):
            if isinstance(obj, ObjectId):
                return str(obj)
            elif isinstance(obj, datetime):
                return obj.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(obj, Binary) or isinstance(obj, bytes):
                return base64.b64encode(obj).decode('utf-8')
            raise TypeError(f'Object of type {type(obj)} is not JSON serializable')

        return json.dumps(obj, default=custom_default, **kwargs)

    def loads(self, s, **kwargs):
        return json.loads(s, **kwargs)

# Then configure all the folders and settings
SCREENSHOTS_FOLDER = os.path.join('static', 'search_screenshots')
os.makedirs(SCREENSHOTS_FOLDER, exist_ok=True)
app.config['SCREENSHOTS_FOLDER'] = SCREENSHOTS_FOLDER

# Rest of your configurations
app.json_provider_class = CustomJSONProvider
app.json = CustomJSONProvider(app)
app.secret_key = os.urandom(24)

# Initialize SocketIO after app is created
socketio = SocketIO(app, cors_allowed_origins="*")

# Add the after_request handler here
@app.after_request
def after_request(response):
    origin = request.headers.get('Origin')
    if origin in ['http://127.0.0.1:3000', 'http://localhost:3000']:
        response.headers.add('Access-Control-Allow-Origin', origin)
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

# Configure folders (keep your existing folder configuration)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# Configure folders
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

# Initialize Google Vision and custom classes
# vision_client = google_cred_setup()
vision_client = vision.ImageAnnotatorClient()
rd_obj = ReceiptDetection()
rf_obj = ReceiptFields()

# Database setup (replace with your connection details)
# MongoDB setup section
try:
    client = MongoClient("mongodb://localhost:27017/")
    db = client["Petition_Ledger"]
    user_collection = db["users"]
    collection = db['petition_data']
    petition_collection = db["petitions"]
    audit_trail_collection = db["audit_trail"]
    scraped_data_collection = db["scraped_data"]  # Add this line
    client.admin.command('ping')
    print("DB connected successfully")
except Exception as e:
    print(f"MongoDB connection failed: {e}")
    # Consider graceful shutdown or fallback behavior

@app.route('/api/latest-petition', methods=['GET'])
def get_latest_petition():
    try:
        template_path = 'Copy of MASTER TEMPLATE.xlsx'
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Find the highest petition number that has been used
            highest_petition = 0
            for r in range(7, 76):
                # Check if row has any data in columns B through G
                has_data = False
                for col in range(2, 8):  # B through G
                    if ws.cell(row=r, column=col).value is not None:
                        has_data = True
                        break
                
                if ws.cell(row=r, column=1).value is not None and has_data:
                    highest_petition = r - 6  # Convert row to petition number
            
            wb.close()
            
            # Next petition should be one more than the highest used
            next_petition = highest_petition + 1
            
            return jsonify({
                'latest_petition': f'petition{next_petition}',
                'status': 'new'
            })

        # Fallback to database check
        latest_petition = db.collection.find_one(
            {'processed': True},  # Only consider fully processed petitions
            sort=[('petition_number', -1)]
        )

        if latest_petition:
            next_num = int(latest_petition['petition_number']) + 1
            return jsonify({
                'latest_petition': f'petition{next_num}',
                'status': 'from_db'
            })
        else:
            return jsonify({
                'latest_petition': 'petition1',
                'status': 'new'
            })

    except Exception as e:
        print(f"Error in get_latest_petition: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')

        if not username or not password:
            return jsonify({'message': 'Username and password are required.'}), 400

        # Find the user in the database
        user = db['users'].find_one({'username': username})

        if user and checkpw(password.encode('utf-8'), user['password'].encode('utf-8') if isinstance(user['password'], str) else user['password']):
            # Remove the ObjectId from the user or serialize it to a string
            user['_id'] = str(user['_id'])
            # Remove the password from the user object before returning it
            del user['password']
            return jsonify({'message': 'Login successful', 'user': user}), 200
        else:
            return jsonify({'message': 'Invalid credentials'}), 401

    except Exception as e:
        print(f"Error in login: {str(e)}")
        return jsonify({'message': 'Server error', 'error': str(e)}), 500


from flask import request, jsonify


@app.route('/api/admin-view', methods=['GET'])
def admin_view():
    try:
        # Simulate the username validation (this would be done with a real session or token in production)
        user = request.args.get('username')
        collection = request.args.get('collection')  # Get the collection parameter

        # Debugging: Print the parameters
        print(f"Received request with username={user} and collection={collection}")

        if not user:
            return jsonify({'message': 'Username not provided'}), 400

        # Fetch the user info from the users collection
        user_info = db['users'].find_one({'username': user})

        if not user_info:
            return jsonify({'message': 'User not found'}), 404

        # Check if the user is an admin
        if user_info.get('usertype') != 'admin':
            return jsonify({'message': 'Access denied. Admins only.'}), 403

        # Handle fetching data based on the collection
        if collection == 'users':
            # Fetch all users, excluding the password and _id fields
            users = list(user_collection.find({}, {'password': 0, '_id': 0}))  # Excluding password and _id
            return jsonify({'users': users}), 200

        elif collection == 'petitions':
            # Remove feedback fields from query
            petitions = list(petition_collection.find({}, {
                'timestamp': 0,
                'screenshot': 0,
                '_id': 0,
                # Remove feedback-related fields
            }))
            return jsonify({'petitions': petitions}), 200

        elif collection == 'audit_trail':
            # Fetch all audit trail data
            audit_trail = list(audit_trail_collection.find({}, {'_id': 0}))  # Excluding _id
            return jsonify({'audit_trail': audit_trail}), 200

        else:
            return jsonify({'message': 'Invalid collection specified.'}), 400

    except Exception as e:
        print(f"Error in admin_view: {str(e)}")
        return jsonify({'error': str(e)}), 500




@app.route('/api/process-petition', methods=['POST'])
def process_petition():
    try:
        print("Received petition request")
        data = request.get_json()
        print("Request data:", data)

        if not data:
            return jsonify({"error": "No data received"}), 400

        required_fields = ['first_name', 'last_name', 'address', 'zip_code', 'petition']
        for field in required_fields:
            if field not in data:
                print(f"Missing field: {field}")
                return jsonify({"error": f"Missing required field: {field}"}), 400

        print("Processing petition with data:", data)
        result = process_petition_text(
            first_name=data['first_name'],
            last_name=data['last_name'],
            address=data['address'],
            zip_code=data['zip_code'],
            petition=data['petition'],
            data={}
        )
        print("Processed result:", result)

        if result is None:
            return jsonify({
                "success": False,
                "error": "No results received from processing"
            }), 500

        return jsonify({
            "success": True,
            "data": result
        })

    except Exception as e:
        print(f"Detailed error in process_petition: {str(e)}")
        import traceback
        print("Traceback:", traceback.format_exc())
        return jsonify({"error": str(e)}), 500
@app.route('/api/process-image', methods=['POST'])
def process_image():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    try:
        # Save uploaded file
        server_url = request.url_root.rstrip('/')
        uploaded_image_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(uploaded_image_path)

        # Rotate and process the uploaded image
        rotated_image = rotate_image_based_on_metadata(PILImage.open(uploaded_image_path))
        c_image = rd_obj.predict(rotated_image)

        if not c_image:
            return jsonify({'error': 'No receipt detected in the image'}), 400

        chunks_data = []
        processed_images = []
        
        # Generate a unique identifier for this batch of images
        batch_id = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Process detected chunks with retry mechanism
        for idx, pil_img in enumerate(c_image.values(), start=1):
            try:
                # Create unique filename for each chunk using batch_id
                chunk_filename = f"img_data_chunk_{batch_id}_{idx}.png"
                chunk_path = os.path.join(app.config['PROCESSED_FOLDER'], chunk_filename)
                pil_img.save(chunk_path)

                # Extract fields and perform OCR
                fields = rf_obj.predict_with_pad(chunk_path, padding=5)
                if not fields:
                    raise Exception("No fields detected")
                    
                fields = dict(sorted(fields.items(), key=lambda item: item[0]))

                imgg_for_ocr = concatenate_pil_images_vertically_with_marginssss(
                    list(fields.values()), margin_percentage=60
                )

                chunk_text = "Not detected"
                
                # Create new Vision client for each request to avoid timeout
                vision_client = vision.ImageAnnotatorClient()
                
                with io.BytesIO() as output:
                    imgg_for_ocr.save(output, format="PNG")
                    content = output.getvalue()

                # Create an Image object
                image = vision.Image(content=content)

                # Define the image context
                image_context = vision.ImageContext(language_hints=["en"])

                # Perform document text detection
                response = vision_client.document_text_detection(
                    image=image,
                    image_context=image_context
                )
                chunk_text = response.full_text_annotation.text.strip() if response.full_text_annotation else "Not detected"
                
                # Extract structured data
                full_name, address, zip_code = extract_info_updated(chunk_text)
                first_name = full_name.split()[0] if full_name and full_name.split() else None
                last_name = full_name.split()[-1] if full_name and full_name.split() else None

                # Update chunk_data with unique screenshot path
                chunk_data = {
                    'chunk_number': idx,
                    'image_url': f'{server_url}/processed_images/{chunk_filename}',
                    'batch_id': batch_id,  # Add batch_id to track related chunks
                    'ocr_data': {
                        'first_name': first_name or "Not detected",
                        'last_name': last_name or "Not detected",
                        'full_name': full_name or "Not detected",
                        'address': address or "Not detected",
                        'zip_code': zip_code or "Not detected",
                        'raw_ocr_text': chunk_text
                    }
                }
                chunks_data.append(chunk_data)
                processed_images.append(pil_img)

            except Exception as e:
                print(f"Error processing chunk {idx}: {str(e)}")
                # Add error handling...

        # Save concatenated image with batch_id
        if processed_images:
            concatenated_filename = f"img_data_combined_{batch_id}.png"
            concatenated_path = os.path.join(app.config['PROCESSED_FOLDER'], concatenated_filename)
            concatenated_image = concatenate_pil_images_vertically_with_marginssss(
                processed_images, margin_percentage=60
            )
            concatenated_image.save(concatenated_path)

        response_data = {
            'message': 'Processing completed',
            'batch_id': batch_id,
            'uploaded_image_url': f'{server_url}/uploads/{file.filename}',
            'chunks': chunks_data,
            'concatenated_image_url': f'{server_url}/processed_images/{concatenated_filename}' if processed_images else None
        }

        return jsonify(response_data), 200

    except Exception as e:
        print(f"Error in process_image: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/processed_images/<path:filename>')
def processed_file(filename):
    return send_from_directory(app.config['PROCESSED_FOLDER'], filename)


import traceback

@app.route('/api/insert-all-automation-data', methods=['POST'])
def insert_all_automation_data():
    try:
        data = request.get_json()
        chunks = data.get('chunks', [])
        petition = data.get('petition')
        batch_id = data.get('batch_id')  # Get batch_id from request
        
        print(f"\n=== Starting Scraping Process for batch {batch_id} ===")
        results = []
        browser = None
        
        try:
            # Initialize browser once
            print("Setting up Chrome WebDriver...")
            options = Options()
            options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            
            service = Service(ChromeDriverManager().install())
            browser = webdriver.Chrome(service=service, options=options)
            print("Chrome WebDriver initialized successfully")
            
            # Authenticate once
            if not authenticate(browser):
                raise Exception("Authentication failed")
            print("Authentication successful")
            
            # Process all chunks with the same browser session
            for chunk in chunks:
                print(f"\nProcessing chunk {chunk['chunk_number']} from batch {batch_id}...")
                try:
                    ocr_data = chunk['ocr_data']
                    
                    # Include batch_id in the data passed to process_petition_text
                    result = process_petition_text(
                        browser=browser,
                        first_name=ocr_data['first_name'],
                        last_name=ocr_data['last_name'],
                        address=ocr_data['address'],
                        zip_code=ocr_data['zip_code'],
                        petition=petition,
                        data={
                            'chunk_number': chunk['chunk_number'],
                            'batch_id': batch_id
                        }
                    )
                    
                    # Add chunk info and batch_id to result
                    result['chunk_image_url'] = chunk['image_url']
                    result['batch_id'] = batch_id
                    results.append(result)
                    
                    # Include batch_id in socket emission
                    socketio.emit('chunk_processed', {
                        'chunk_number': chunk['chunk_number'],
                        'batch_id': batch_id,
                        'result': result
                    })
                    
                except Exception as chunk_error:
                    print(f"Error processing chunk {chunk['chunk_number']}: {chunk_error}")
                    error_result = {
                        'chunk_number': chunk['chunk_number'],
                        'batch_id': batch_id,
                        'success': False,
                        'message': str(chunk_error),
                        'found': False
                    }
                    results.append(error_result)
                    socketio.emit('chunk_processed', {
                        'chunk_number': chunk['chunk_number'],
                        'batch_id': batch_id,
                        'result': error_result
                    })
                    
        finally:
            if browser:
                browser.quit()
        
        return jsonify({
            "success": True,
            "batch_id": batch_id,
            "results": results
        })
        
    except Exception as e:
        print(f"Error in batch processing: {e}")
        return jsonify({
            "error": str(e),
            "batch_id": batch_id
        }), 500


@app.route('/api/signup', methods=['POST'])
def signup():
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        usertype = data.get('usertype')  # Accept usertype from frontend

        if not username or not password or not usertype:
            return jsonify({'message': 'Username, password, and usertype are required.'}), 400

        # Check if the user already exists
        if db['users'].find_one({'username': username}):
            return jsonify({'message': 'User already exists'}), 409

        # Hash the password
        hashed_password = hashpw(password.encode('utf-8'), gensalt())

        # Store the user with usertype
        db['users'].insert_one({
            'username': username,
            'password': hashed_password.decode('utf-8'),  # Store as string
            'usertype': usertype  # Add usertype to the document
        })

        return jsonify({'message': 'Signup successful'}), 201

    except Exception as e:
        print(f"Error in signup: {str(e)}")
        return jsonify({'message': 'Server error', 'error': str(e)}), 500

# Set custom encoder for Flask
@app.route("/api/search", methods=["GET"])
def search():
    try:
        params = {k: v.strip() for k, v in request.args.to_dict().items() if v.strip()}
        search_method = params.pop('method', None)
        
        query = {}
        
        if search_method == "method1":  # First Initial + Last Initial
            if 'first_name' in params:
                query['first_name'] = {"$regex": f"^{params['first_name']}", "$options": "i"}
            if 'last_initial' in params:
                query['last_name'] = {"$regex": f"^{params['last_initial']}", "$options": "i"}
            if 'street_number' in params:
                query['address'] = {"$regex": f"^{params['street_number']}", "$options": "i"}
            if 'street_initial' in params:
                query['address'] = {"$regex": f"{params['street_number']}.*{params['street_initial']}", "$options": "i"}
            if 'zip_code' in params:
                query['zip_code'] = params['zip_code']
        elif search_method == "method2":  # First Initial + Full Last Name
            # Similar pattern for other methods...
            pass
        else:  # Default search (all fields)
            if 'first_name' in params:
                query['first_name'] = {"$regex": f"^{params['first_name']}", "$options": "i"}
            if 'last_name' in params:
                query['last_name'] = {"$regex": f"^{params['last_name']}", "$options": "i"}
            if 'address' in params:
                query['address'] = {"$regex": f"^{params['address']}", "$options": "i"}
            if 'zip_code' in params:
                query['zip_code'] = params['zip_code']
            
        # Perform search with the constructed query
        if query:
            results = list(collection.find(query, {
                "_id": 0,
                "first_name": 1,
                "last_name": 1,
                "address": 1,
                "zip_code": 1
            }))
            
            # Remove duplicates
            unique_results = []
            seen = set()
            for result in results:
                key = f"{result['first_name']}-{result['last_name']}-{result['address']}-{result['zip_code']}"
                if key not in seen:
                    seen.add(key)
                    unique_results.append(result)
            
            return jsonify({"results": unique_results})
        else:
            return jsonify({"results": []})

    except Exception as e:
        print(f"Search error: {str(e)}")
        return jsonify({"error": str(e)}), 500

def get_next_petition_number(ws):
    """Find the next available petition number"""
    last_petition = 0
    for row in range(7, 76):  # Check rows 7-75
        petition_num = ws.cell(row=row, column=1).value
        if isinstance(petition_num, (int, float)) and petition_num > last_petition:
            last_petition = int(petition_num)
    return last_petition + 1

@app.route('/api/get-next-petition', methods=['GET'])
def get_next_petition():
    try:
        template_path = 'Copy of MASTER TEMPLATE.xlsx'
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'error': 'Template file not found'}), 404
            
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Find next available row and petition number
        next_petition = None
        for r in range(7, 76):
            if ws.cell(row=r, column=1).value is None:
                next_petition = (r - 6)  # Row 7 = Petition 1, Row 8 = Petition 2, etc.
                break
        
        wb.close()
        
        if next_petition is None:
            return jsonify({'success': False, 'error': 'No available petition slots'}), 400
            
        return jsonify({
            'success': True,
            'nextPetition': next_petition
        })
        
    except Exception as e:
        print(f"Error getting next petition: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.errorhandler(Exception)
def handle_error(error):
    print(f"Error: {str(error)}")
    return jsonify({'error': str(error)}), 500

@app.route('/search_screenshots/<path:filename>')
def serve_screenshot(filename):
    return send_from_directory(app.config['SCREENSHOTS_FOLDER'], filename)

@app.route('/api/process-single-signature', methods=['POST'])
def process_single_signature():
    browser = None
    try:
        data = request.json
        print(f"\nProcessing signature {data.get('chunk_number')}:")
        
        # Initialize browser first
        print("Setting up Chrome WebDriver...")
        options = Options()
        options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        service = Service(ChromeDriverManager().install())
        browser = webdriver.Chrome(service=service, options=options)
        print("Chrome WebDriver initialized successfully")
        
        # Authenticate
        if not authenticate(browser):
            raise Exception("Authentication failed")
        print("Authentication successful")
        
        # Get the search method
        search_method = data.get('method', '')
        
        if search_method == 'simple':
            # Simple search: Use the fields exactly as entered
            first_name = data.get('first_name', '')
            last_name = data.get('last_name', '')
            address = data.get('address', '')
            zip_code = data.get('zip_code', '')
            
            print(f"Simple Search Parameters:")
            print(f"First Name: {first_name}")
            print(f"Last Name: {last_name}")
            print(f"Address: {address}")
            print(f"Zip: {zip_code}")
            
            # Process the signature with the exact parameters
            result = process_petition_text(
                browser=browser,  # Pass the browser instance
                first_name=first_name,
                last_name=last_name,
                address=address,
                zip_code=zip_code,
                petition=data.get('petition'),
                data={
                    'chunk_number': data.get('chunk_number'),
                    'method': search_method
                }
            )
            
            return jsonify(result)
            
        elif search_method == 'streetCentric':
            # Street Centric: First 3 letters of names, street number, first letter of street
            first_name = data.get('first_name', '')[:3]
            last_name = data.get('last_name', '')[:3]
            address = data.get('address', '')
            zip_code = data.get('zip_code', '')
            
            result = process_petition_text(
                browser=browser,  # Pass the browser instance
                first_name=first_name,
                last_name=last_name,
                address=address,
                zip_code=zip_code,
                petition=data.get('petition'),
                data={
                    'chunk_number': data.get('chunk_number'),
                    'method': search_method
                }
            )
            
            return jsonify(result)
            
        elif search_method == 'nameCentric':
            # Name Centric: First letter of names, street number, first 3 letters of street
            first_name = data.get('first_name', '')[0] if data.get('first_name') else ''
            last_name = data.get('last_name', '')[0] if data.get('last_name') else ''
            address = data.get('address', '')
            zip_code = data.get('zip_code', '')
            
            result = process_petition_text(
                browser=browser,  # Pass the browser instance
                first_name=first_name,
                last_name=last_name,
                address=address,
                zip_code=zip_code,
                petition=data.get('petition'),
                data={
                    'chunk_number': data.get('chunk_number'),
                    'method': search_method
                }
            )
            
            return jsonify(result)
        
        else:
            raise ValueError(f"Invalid search method: {search_method}")

    except Exception as e:
        print(f"Error processing signature: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "chunk_number": data.get('chunk_number'),
            "found": False
        }), 500
        
    finally:
        if browser:
            try:
                browser.quit()
            except:
                pass

@app.route('/api/update-spreadsheet', methods=['POST'])
def update_spreadsheet():
    try:
        data = request.json
        value = data.get('value')
        chunk_number = int(data.get('chunkNumber', 1))
        petition = data.get('petition')
        
        if not petition:
            return jsonify({
                'success': False, 
                'error': 'No petition specified'
            }), 400
            
        template_path = 'Copy of MASTER TEMPLATE.xlsx'
        petition_number = int(petition.replace('petition', ''))
        
        # Retry mechanism for file access
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                wb = load_workbook(template_path)
                ws = wb.active
                
                # Calculate row (petition1 -> row 7, petition2 -> row 8, etc.)
                row = petition_number + 6
                
                # Initialize row if needed
                if not ws.cell(row=row, column=1).value:
                    ws.cell(row=row, column=1, value=petition_number)
                    # Clear existing data in the row
                    for col in range(2, 19):  # B through S
                        ws.cell(row=row, column=col).value = None
                
                # Map chunk number to correct column (chunk1 -> B, chunk2 -> C, etc.)
                col = chunk_number + 1  # B is column 2
                
                # Validate column range
                if col < 2 or col > 10:  # B through J
                    return jsonify({
                        'success': False, 
                        'error': f'Invalid chunk number. Only chunks 1-9 allowed'
                    }), 400
                
                # Update the signature cell
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Update formulas for the row
                formulas = {
                    11: f'=SUMPRODUCT(--(B{row}:J{row}=1),1,--(B{row}:J{row}=0.1),0.1)',  # K: Total valid (fixed)
                    12: f'=COUNTIF(B{row}:J{row},"DUP")',  # L: Duplicates
                    13: f'=COUNTIF(B{row}:J{row},"X")',    # M: Purges
                    14: f'=L{row}+M{row}',                 # N: Total purge
                    16: f'=COUNTIF(B{row}:J{row},"V")',    # P: Voter validations
                    17: f'=COUNTA(B{row}:J{row})',         # Q: Total checked
                    18: f'=COUNTIF(B{row}:J{row},1)',      # R: Good signatures
                    19: f'=COUNTIF(B{row}:J{row},0.1)'     # S: Bad signatures
                }
                
                # Apply formulas
                for col_num, formula in formulas.items():
                    ws.cell(row=row, column=col_num).value = formula
                
                # Save changes
                wb.save(template_path)
                
                # Calculate current totals for response
                totals = {
                    'validCount': sum(1 for c in range(2, 11) if ws.cell(row=row, column=c).value in [1, 0.1]),
                    'duplicateCount': sum(1 for c in range(2, 11) if ws.cell(row=row, column=c).value == 'DUP'),
                    'purgeCount': sum(1 for c in range(2, 11) if ws.cell(row=row, column=c).value == 'X'),
                    'voterValidated': sum(1 for c in range(2, 11) if ws.cell(row=row, column=c).value == 'V'),
                    'totalChecked': sum(1 for c in range(2, 11) if ws.cell(row=row, column=c).value is not None)
                }
                
                return jsonify({
                    'success': True,
                    'message': f'Updated cell {chr(65+col)}{row} with value {value}',
                    'petition': petition,
                    'chunk': chunk_number,
                    'totals': totals
                })
                
            except PermissionError:
                retry_count += 1
                if retry_count == max_retries:
                    return jsonify({
                        'success': False,
                        'error': 'Excel file is locked. Please close it and try again.'
                    }), 423
                time.sleep(1)
                
            finally:
                if 'wb' in locals():
                    try:
                        wb.close()
                    except:
                        pass
                    
    except Exception as e:
        print(f"Error updating spreadsheet: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    print("Starting Flask server...")
    print("Access the API at: http://127.0.0.1:5000")
    try:
        socketio.run(
            app,
            debug=True,
            host='127.0.0.1',
            port=5000,
            use_reloader=False  # Disable auto-reloader
        )
    except KeyboardInterrupt:
        print("\nShutting down gracefully...")
    except Exception as e:
        print(f"\nError occurred: {e}")
    finally:
        print("Server stopped")